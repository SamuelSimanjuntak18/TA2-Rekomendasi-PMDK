from flask import Flask, render_template, request, redirect, url_for, send_from_directory, jsonify
from werkzeug.utils import secure_filename
import pandas as pd
import joblib
from openpyxl import load_workbook
import pdfkit
app = Flask(__name__)

# Import model yang sudah di TRAINING
model = joblib.load('D:\\SEMESTER7&8\\TA1\\Referensi\\TRIAL\\TA2-Rekomendasi-PMDK\\Model\\ModelTA.pkl')
# Mapping jurusan
mapping_prodi = {
    'DIII Teknologi Informasi': 1,
    'DIII Teknologi Komputer': 2,
    'DIV Teknologi Rekayasa Perangkat Lunak': 3,
    'S1 Teknik Informatika': 4,
    'S1 Manajemen Rekayasa': 5,
    'S1 Sistem Informasi': 6,
    'S1 Teknik Bioproses': 7,
    'S1 Teknik Elektro': 8
}
# Fungsi konversi HTML ke PDF
def html_to_pdf(html_content, output_path):
    options = {
        'page-size': 'Letter',
        'margin-top': '0.75in',
        'margin-right': '0.75in',
        'margin-bottom': '0.75in',
        'margin-left': '0.75in',
    }
    pdfkit.from_string(html_content, output_path, options=options)

# Halaman awal
@app.route('/')
def index():
    return render_template('index.html')

# Upload file
@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        file.save(filename)

        try:
            if filename.endswith('.xlsx'):
                workbook = load_workbook(filename=filename)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                columns = data[0]
                data_df = pd.DataFrame(data[1:], columns=columns)

            elif filename.endswith('.csv'):
                data_df = pd.read_csv(filename)

            else:
                raise ValueError("File harus memiliki ekstensi (.xlsx) atau CSV (.csv) ")

            # Tambahkan kolom "Nama" dengan nilai default kosong jika tidak ada
            if 'Nama' not in data_df.columns:
                data_df['Nama'] = ''
                show_nama = False
            else:
                show_nama = True

            # Tambahkan kolom "Nomor" berisi nomor urut jika belum ada
            if 'Nomor' not in data_df.columns:
                data_df.insert(0, 'Nomor', range(1, len(data_df) + 1))

            # Ambil Mapping Jurusan
            data_df[['Pilihan 1', 'Pilihan 2', 'Pilihan 3']] = data_df[['Pilihan 1', 'Pilihan 2', 'Pilihan 3']].replace(mapping_prodi)
            # Buat Prediksi
            predictions = model.predict(data_df.drop(columns=['Nama', 'Nomor']))  # Drop kolom 'Nama' dan 'Nomor' sebelum membuat prediksi
            probabilities = model.predict_proba(data_df.drop(columns=['Nama', 'Nomor']))

            # Data Frame untuk Prediksi Kelulusan
            result_df = pd.DataFrame({
                'Nomor': data_df['Nomor'],
                'Nama': data_df['Nama'] if 'Nama' in data_df.columns else None,  # Sertakan kolom "Nama" jika ada
                'Pilihan 1': [f"Lulus" if pred == 1 else f"Tidak Lulus" for pred, prob in zip(predictions, probabilities[:, 1])],
                'Pilihan 2': [f"Lulus" if pred == 1 else f"Tidak Lulus" for pred, prob in zip(predictions, probabilities[:, 1])],
                'Pilihan 3': [f"Lulus " if pred == 1 else f"Tidak Lulus" for pred, prob in zip(predictions, probabilities[:, 1])]
        })
                #  'Pilihan 1': [f"Lulus (Prob: {prob:.10f})" if pred == 1 else f"Tidak Lulus (Prob: {prob:.10f})" for pred, prob in zip(predictions, probabilities[:, 1])],
                # 'Pilihan 2': [f"Lulus (Prob: {prob:.10f})" if pred == 1 else f"Tidak Lulus (Prob: {prob:.10f})" for pred, prob in zip(predictions, probabilities[:, 1])],
                # 'Pilihan 3': [f"Lulus (Prob: {prob:.10f})" if pred == 1 else f"Tidak Lulus (Prob: {prob:.10f})" for pred, prob in zip(predictions, probabilities[:, 1])]       
            
            if not show_nama:
                result_df = result_df.drop(columns=['Nama'])  

            return render_template('result.html', result=result_df.to_html(index=False))

        except Exception as e:
            return str(e)

# Halaman hasil
@app.route('/result')
def result():
    return render_template('result.html')

# Fungsi untuk mengunduh PDF
@app.route('/download_pdf', methods=['POST'])
def download_pdf():
    try:
        # Periksa apakah ada file HTML yang dikirimkan dalam permintaan POST
        if 'html_content' not in request.form:
            raise ValueError("HTML content not found in the request")

        html_content = request.form['html_content']  # Ambil konten HTML dari formulir
        pdf_filename = 'Hasil_Rekomendasi.pdf'  # Nama file PDF yang akan diunduh
        pdf_path = 'static/' + pdf_filename  # Path untuk menyimpan file PDF

        html_to_pdf(html_content, pdf_path)  # Konversi HTML ke PDF
        return jsonify({'success': True, 'pdf_filename': pdf_filename})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# Rute untuk mengunduh file PDF
@app.route('/download/<path:filename>', methods=['GET'])
def download(filename):
    directory = 'static'
    return send_from_directory(directory, filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=False, host="0.0.0.0")
